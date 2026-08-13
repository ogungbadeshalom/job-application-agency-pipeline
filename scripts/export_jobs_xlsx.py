#!/usr/bin/env python3
"""EXPIRIMENTAL — export scraped jobs to a spreadsheet (no DB insert).

Usage:
    python scripts/export_jobs_xlsx.py <config_json> <jobs_json> <out.xlsx>

Reads the jobs JSON produced by run_jobspy.py and writes an .xlsx for the admin
to inspect what a scrape actually returns. This lets you audit the pipeline
without polluting the candidate's queue. May be removed later.
"""
import sys, json
import pandas as pd

if len(sys.argv) < 4:
    print("Usage: export_jobs_xlsx.py <config_json> <jobs_json> <out.xlsx>", file=sys.stderr)
    sys.exit(2)

config = json.loads(sys.argv[1])
with open(sys.argv[2], 'r', encoding='utf-8') as f:
    jobs_file = f.read()
jobs = json.loads(jobs_file)
out_file = sys.argv[3]

def cell(v):
    if isinstance(v, float) and (v != v):   # NaN
        return None
    return v

rows = []
for j in jobs:
    site = cell(j.get('site')) or ''
    url = cell(j.get('job_url'))
    rows.append({
        'title': cell(j.get('title')),
        'company': cell(j.get('company')),
        'site': site,
        'link': url,
        'url': url,                       # kept as plain text fallback + used for hyperlink
        'comp_min': cell(j.get('interval_amount')),
        'currency': cell(j.get('currency')),
        'location': cell(j.get('location')),
        'date_posted': cell(j.get('date_posted')),
        'description': (lambda s: (s.strip() if isinstance(s, str) else None) or None)(cell(j.get('description'))),
    })

# Headline summary
meta = {
    'search_terms': ', '.join(config.get('search_terms', [])),
    'sites': ', '.join(config.get('sites', [])),
    'location': config.get('location', ''),
    'remote_only': config.get('remote_only'),
    'jobs_scraped': len(rows),
}

import pandas as pd
with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
    pd.DataFrame([meta]).T.rename(columns={0: 'value'}).to_excel(writer, sheet_name='Overview', header=False)
    pd.DataFrame(rows).to_excel(writer, sheet_name='Jobs', index=False)

# Post-process: make the 'link' cell a clickable hyperlink labeled "Open (<site>)"
from openpyxl import load_workbook
wb = load_workbook(out_file)
ws = wb['Jobs']
# find the link column index (header 'link')
hdr = [c.value for c in ws[1][:ws.max_column]]
try:
    li = hdr.index('link')
except ValueError:
    li = None
if li is not None:
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=li + 1)
        if c.value:
            site_txt = ws.cell(row=r, column=hdr.index('site') + 1).value or ''
            c.hyperlink = c.value
            c.style = 'Hyperlink'
            c.value = f"Open ({site_txt})" if site_txt else "Open"
wb.save(out_file)

print(f"[ok] {len(rows)} jobs written to {out_file}", file=sys.stderr)